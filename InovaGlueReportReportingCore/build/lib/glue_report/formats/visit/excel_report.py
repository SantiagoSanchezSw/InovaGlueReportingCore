from openpyxl import load_workbook
from django.conf import settings
import uuid
import re
# Models
from inova_core.models import Fiscalizacion, Cliente
from reporting_core.models import CustomReport
# Classes
from reporting_core.classes.inova_report import InOvaReport
from reporting_core.classes.aws import AWSClient
# Helpers
from reporting_core.helpers.file import remove_generated_file
# Types
from reporting_core.types import GeneratedReport
from typing import Optional, Pattern, List
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell
# Enums
from reporting_core.enums import Extension
from inova_base.helpers.user.client import get_client_language


class ExcelReport(InOvaReport):
    visit: Fiscalizacion
    template: CustomReport
    preview: bool = False
    client: Cliente
    question_regex: Pattern
    basic_info_regex: Pattern
    workbook: Workbook
    worksheet: Worksheet
    file_id: uuid.UUID
    generated_files: List[str]

    def __init__(self, template: CustomReport, client: Cliente, visit: Optional[Fiscalizacion] = None):
        if visit:
            self.visit = visit
        else:
            self.preview = True
            self.visit = Fiscalizacion.objects.using(settings.REPORTING_CORE_DB).filter(
                norma__pk=template.rule_id
            ).select_related(
                'entidad', 'entidad__productor', 'entidad__productor__persona', 'profesional',
                'profesional__usuario', 'profesional__usuario__persona'
            ).last()
        self.client = client
        self.template = template
        self.question_regex = re.compile(r'\[\[qID([0-9-?]+)]]')
        self.basic_info_regex = re.compile(r'\[\[(.+?)]]')
        self.generated_files = []

    def generate(self) -> GeneratedReport:
        self.language = get_client_language(self.client, using=settings.REPORTING_CORE_DB)
        self.file_id = uuid.uuid4()
        clean_workbook_path = self.download_file_to_tmp_folder(
            file_url=self.template.temp_excel_template,
            file_name=str(self.file_id),
            extension=Extension.XLSX
        )
        self.generated_files.append(clean_workbook_path)
        self.workbook = load_workbook(clean_workbook_path)
        processed_workbook_path = '{}processed_{}.xlsx'.format(settings.TEMPORAL_DIR, self.file_id)
        self.process_workbook()
        self.workbook.save(processed_workbook_path)
        self.generated_files.append(processed_workbook_path)
        exported_report_path = self.export_excel_to_pdf(processed_workbook_path)
        self.remove_generated_files()
        return GeneratedReport(
            path=exported_report_path,
            file_name='exported_{}.pdf'.format(self.file_id),
            custom_report=self.template
        )

    def process_workbook(self) -> None:
        self.worksheet = self.workbook.active
        for row in self.worksheet.iter_rows():
            for cell in row:
                if cell.value:
                    self.handle_worksheet_cell(cell)
        return None

    def handle_worksheet_cell(self, cell: Cell) -> None:
        cell_value = str(cell.value).strip()
        replace_value = self.get_replace_value(cell_value)
        if replace_value is not None:
            if isinstance(replace_value, list):
                row_number = cell.row
                for i in replace_value:
                    self.worksheet[cell.column_letter + str(row_number)] = i
                    row_number = row_number + 1
            elif isinstance(replace_value, str):
                cell.value = replace_value
        return None

    def get_replace_value(self, cell_value: str):
        if self.is_question_key(cell_value):
            key = cell_value.replace('[[qID', '').replace(']]', '')
            values = key.split('-')
            return self.replace_question_key(values)
        elif self.is_basic_info_key(cell_value):
            return self.replace_basic_info_key(cell_value)
        return None

    def replace_question_key(self, key_values: List[str]) -> str:
        if len(key_values) == 1:
            return self.get_answer_in_chapter(self.visit.entidad.informacion_general, int(key_values[0]))
        else:
            visit_chapter = self.get_item('capitulo', 'pk', int(key_values[0]), self.visit.capitulos)
            if visit_chapter:
                if len(key_values) == 3:
                    return self.handle_compound_replace(key_values, visit_chapter)
                else:
                    return self.get_answer_in_chapter(visit_chapter['respuestas'], int(key_values[1]))
        return ''

    def replace_basic_info_key(self, cell_value: str) -> str:
        blocks = self.get_basic_info_blocks()
        if cell_value in blocks.keys():
            return blocks[cell_value]()
        return ''

    def handle_compound_replace(self, key_values, visit_chapter):
        compound_values = []
        compound = self.get_answer_in_chapter(visit_chapter['respuestas'], int(key_values[1]))
        for compound_item in compound:
            compound_values.append(
                self.get_answer_in_chapter(list(compound_item['items']),
                                           int(key_values[2]),
                                           question_key='subpregunta')
            )
        return compound_values

    def export_excel_to_pdf(self, path: str) -> str:
        upload_path = '{}/{}.{}'.format(settings.ENVIRONMENT, self.file_id, Extension.XLSX)
        output_path = '{}/{}.{}'.format(settings.ENVIRONMENT, self.file_id, Extension.PDF)
        s3_client = AWSClient('s3')
        s3_client.upload_file(file_path=path, bucket=settings.UNOCONV_BUCKET_NAME, upload_to=upload_path)
        AWSClient(service='lambda').invoke_function(
            function_name=settings.UNOCONV_FUNCTION_NAME,
            invocation_type='RequestResponse',
            payload={'input_path': upload_path, 'output_path': output_path}
        )
        export_path = self.download_file_to_tmp_folder(
            file_url=settings.UNOCONV_BUCKET_HOST + output_path,
            file_name='exported_{}'.format(self.file_id),
            extension=Extension.PDF
        )
        s3_client.delete_file(file_path=upload_path, bucket=settings.UNOCONV_BUCKET_NAME)
        s3_client.delete_file(file_path=output_path, bucket=settings.UNOCONV_BUCKET_NAME)
        return export_path

    def is_question_key(self, value: str) -> bool:
        return not not self.question_regex.match(value)

    def is_basic_info_key(self, value: str) -> bool:
        return not not self.basic_info_regex.match(value)

    def remove_generated_files(self) -> None:
        for path in self.generated_files:
            remove_generated_file(path)
        return None
